import pandas
import json
from datetime import date
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_match_json(file_name, match_id, processed_match):
    output_file_path = f"results/{(file_name, match_id)[match_id != '']}_data.json"
    with open(output_file_path, "w") as output_file:
        json.dump(processed_match, output_file)       

def build_cumulative_reports(matches, player_data):
    with pandas.ExcelWriter(f"results/Stats_Report_{date.today()}.xlsx", engine="openpyxl", mode="w") as writer:        
        for i, match in enumerate(matches):
            sheet_name = "Match Report"
            if sheet_name in writer.sheets:
                startrow = writer.sheets[sheet_name].max_row
                header = False
                min_row = startrow + 1
            else:
                startrow = 0
                header = True
                min_row = 2

            pandas.DataFrame(match["participants"]).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=header)
            worksheet = writer.sheets[sheet_name]
            if i % 2 == 0:  # apply grey fill to every other match
                grey_fill = PatternFill(start_color="d3d3d3", end_color="d3d3d3", fill_type="solid")
                for row in worksheet.iter_rows(min_row):
                    for cell in row:
                        cell.fill = grey_fill
            thin_border = Border(bottom=Side(style='thin'))
            for row in worksheet.iter_rows():
                if (row[0].row - 1) % 5 == 0 and row[0].row > 1:
                    for cell in row:
                        cell.border = thin_border

        pandas.DataFrame(player_data).T.to_excel(writer, sheet_name="Player Stats", index=True)   
        format_workbook(writer)

def build_match_sheets(matches):
    with pandas.ExcelWriter(f"results/Matches_{date.today()}.xlsx", engine="openpyxl", mode="w") as writer:
        for match in matches:
            sheet_name = f"{match['participants'][0]['player']} vs {match['participants'][5]['player']}"
            pandas.DataFrame(match["participants"]).to_excel(writer, sheet_name=sheet_name, index=False)
        format_workbook(writer)


def build_player_sheets(matches):
     with pandas.ExcelWriter(f"results/Players_{date.today()}.xlsx", engine="openpyxl", mode="w") as writer:
        for match in matches:
            for participant in match["participants"]:
                sheet_name = f"{participant['player']}"
                if sheet_name in writer.sheets:
                    startrow = writer.sheets[sheet_name].max_row
                    df = pandas.DataFrame(participant, index=[0])
                    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
                else:
                    pandas.DataFrame(participant, index=[0]).to_excel(writer, sheet_name=sheet_name, index=False)
        format_workbook(writer)


def format_workbook(writer):
    # Adjust column widths               
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        #Hide puuid column
        worksheet.column_dimensions['A'].hidden = True